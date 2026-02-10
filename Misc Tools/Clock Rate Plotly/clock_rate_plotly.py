from datetime import datetime, timedelta, date
import plotly.graph_objects as go
import plotly.io as pio
from pathlib import Path
import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chartsheet.publish import WebPublishItem

WebPublishItem.sourceRef.expected_type = (str, type(None))
pio.renderers.default= "browser"

def parse_excel_to_lists(path, sheet_name=None):
    wb = load_workbook(path, data_only=True)

    real_sheets = {
        ws.title: ws
        for ws in wb.worksheets
        if isinstance(ws, Worksheet)
    }

    if not real_sheets:
        raise ValueError("Workbook contains no real worksheets (only chartsheets).")

    if sheet_name:
        if sheet_name not in real_sheets:
            raise ValueError(
                f"'{sheet_name}' is not a real worksheet. "
                f"Available worksheets: {list(real_sheets.keys())}"
            )
        ws = real_sheets[sheet_name]
    else:
        ws = list(real_sheets.values())[0]

    headers = [cell.value for cell in ws[1]]
    data = {h: [] for h in headers}

    for row in ws.iter_rows(min_row=2, values_only=True):
        for header, value in zip(headers, row):
            data[header].append(value)

    return data


def correct_data(dates, values):
    corrected_dates, corrected_values = [], []

    for date, value in zip(dates, values):
        try:
            new_date = datetime.strptime(date, "%Y:%j:%H:%M:%S.%f")
            new_value = value
        except (ValueError, TypeError):
            new_date = None
            new_value = None

        corrected_dates.append(new_date)
        corrected_values.append(new_value)

    return corrected_dates, corrected_values


def moving_average_np(values, window):
    values = np.array(values, dtype=float)
    ma = np.convolve(values, np.ones(window)/window, mode='valid')
    pad = [None] * (window - 1)
    return pad + ma.tolist()


def clean_float_list(values):
    cleaned = []
    for v in values:
        try:
            cleaned.append(float(v))
        except (TypeError, ValueError):
            cleaned.append(None)
    return cleaned


def get_y_ranges(plot, duration= None, start_date= None):
    low_value, high_value, offset= 1,-1, 0

    for date, value in zip(plot.dates, plot.values):
        try:
            # Check if in date range:
            if duration == "mission":
                low_range= datetime(1999,7,1).date()
                offset= 5e-10
            elif duration == "biannual" and start_date is not None:
                low_range= datetime.strptime(start_date, "%Y:%j").date()
                offset= 2e-11
            else:
                low_range= plot.start_range
                offset= 2e-11

            if low_range <= date.date() <= plot.end_range:
                if value <= low_value:
                    low_value= value
                if value >= high_value:
                    high_value= value
        except (AttributeError, TypeError):
            pass

    return [low_value - offset, high_value + offset]


def mission_plot(plot):
    "Format and write the mission plot"
    plot.plot.update_layout(
        title= dict(text= f"Mission VCDU Clock Rate", xanchor= "center",
                    yanchor= "top", y= 0.95, x=0.5, font= dict(size= 30)),
        xaxis= dict(showline= True, linewidth= 1, linecolor= "black", mirror= True,
                    title= dict(text= "Date", font= dict(size= 30)),
                    range= [datetime(1999,7,1).date(),
                            plot.end_range + timedelta(days=30)]),
        yaxis= dict(gridcolor= "black", mirror= True, linewidth= 1, linecolor= "black",
                    title= dict(text= "Rate (seconds/seconds)", font= dict(size= 30)),
                    range= get_y_ranges(plot, "mission")))

    # Write to Clock_Timing directory
    set_dir= Path("//noodle/FOT/engineering/ccdm/Clock_Timing/Clock Rate Trending_files")
    plot.plot.write_html(f"{set_dir}/Mission_VCDU_Clock_Rate_Plotly.html")
    plot.plot.write_image(f"{set_dir}/Mission_VCDU_Clock_Rate.png",
                          height=1000, width= 1600, scale= 2)
    print(f" - Mission_VCDU_Clock_Rate.html written to {set_dir}")
    print(f" - Mission_VCDU_Clock_Rate.png written to {set_dir}")

    # Write to vweb directory
    set_dir= Path("//noodle/vweb/fot_web/eng/subsystems/ccdm/Clock_Rate/images")
    plot.plot.write_image(f"{set_dir}/Mission_VCDU_Clock_Rate.png",
                          height=1000, width= 1600, scale= 2)
    print(f" - Mission_VCDU_Clock_Rate.png written to {set_dir}")


def year_plot(plot):
    "Format and write the current year plot"
    plot.plot.update_layout(
        title= dict(xanchor= "center", yanchor= "top", y= 0.95, x=0.5, font= dict(size= 30),
                    text= f"{datetime.now().year} VCDU Clock Rate"),
        xaxis= dict(showline= True, linewidth= 1, linecolor= "black", mirror= True,
                    title= dict(text= "Date", font= dict(size= 30)),
                    range= [plot.start_range - timedelta(days= 5),
                            plot.end_range + timedelta(days=10)]),
        yaxis= dict(gridcolor= "black", mirror= True, linewidth= 1, linecolor= "black",
                    title= dict(text= "Rate (seconds/seconds)", font= dict(size= 30)),
                    range= get_y_ranges(plot)))

    # Write to Clock_Timing directory
    set_dir= Path("//noodle/FOT/engineering/ccdm/Clock_Timing/Clock Rate Trending_files")
    plot.plot.write_image(f"{set_dir}/{datetime.now().year}_Daily_Clock_Rate.png",
                          height=1000, width= 1600, scale= 2)
    print(f" - {datetime.now().year}_Daily_Clock_Rate.png written to {set_dir}")

    # Write to vweb directory
    set_dir= Path("//noodle/vweb/fot_web/eng/subsystems/ccdm/Clock_Rate/images")
    plot.plot.write_image(f"{set_dir}/{datetime.now().year}_Daily_Clock_Rate.png",
                          height=1000, width= 1600, scale= 2)
    print(f" - {datetime.now().year}_Daily_Clock_Rate.png written to {set_dir}")


def biannual_plot(plot):
    "format and write the current biannual plot"
    while True:
        try:
            start_date= datetime.strptime(input(" - Enter biannual period start (yyyy:ddd): "),
                                          "%Y:%j")
            end_date=   datetime.strptime(input(" - Enter biannual period end (yyyy:ddd): "),
                                          "%Y:%j")
        except ValueError:
            print(" - Invalid date format. Please enter dates in 'yyyy:ddd' format.")
            continue
        break

    plot.plot.update_layout(
        title= dict(xanchor= "center", yanchor= "top", y= 0.95, x=0.5, font= dict(size= 30),
                    text= f"{start_date.strftime("%b %Y")} - {end_date.strftime("%b %Y")} Daily Clock Rate"),
        xaxis= dict(showline= True, linewidth= 1, linecolor= "black", mirror= True,
                    title= dict(text= "Date", font= dict(size= 30)),
                    range= [start_date.date() - timedelta(days=5), end_date.date() + timedelta(days=5)]),
        yaxis= dict(gridcolor= "black", mirror= True, linewidth= 1, linecolor= "black",
                    title= dict(text= "Rate (seconds/seconds)", font= dict(size= 30)),
                    range= get_y_ranges(plot, "biannual", start_date.strftime("%Y:%j"))))

    # Write to user desktop directory
    set_dir= Path.home() / "Desktop"
    plot.plot.write_image(f"{set_dir}/Biannual_Daily_Clock_Rate.png",
                          height=1000, width= 1600, scale= 2)
    print(f" - Biannual_Daily_Clock_Rate.png written to {set_dir}")

def generate_plot(self):
    self.dates, self.values = correct_data(self.parsed["RefTime(UTC)"], self.parsed["1-day rate"])
    self.ma07 = moving_average_np(clean_float_list(self.values), window=7)

    plot_obj = go.Figure()

    plot_obj.add_trace(go.Scatter(
        x= self.dates, y= self.values,
        mode="markers", name="Daily Clock Rate"))

    plot_obj.add_trace(go.Scatter(
        x= self.dates, y= self.ma07, mode="lines",
        name="7-Day Moving Average", line= dict(width=3, dash="solid", color="red")))

    plot_obj.add_annotation(
        x= 0.5, y= 1.04, xref= "paper", yref= "paper",
        text= f"(Generated On {datetime.now().date()})", showarrow= False,
        font= dict(size= 18, color= "black"), xanchor= "center")

    # Set General Layout
    plot_obj.update_layout(
        hovermode= "x unified", paper_bgcolor= "rgb(255,255,255)", plot_bgcolor= "rgb(255,255,255)",
        legend= dict(orientation= "h", yanchor= "bottom", y=-0.15, xanchor= "center",
                     x= 0.5, bordercolor= "Black", borderwidth=1))

    return plot_obj


class ClockPlot:

    def __init__(self):
        super().__init__()
        parse_file= Path("N:/engineering/ccdm/Clock_Timing/Clock Rate Trending_files"
                         "/Clock Rate Trending (Data Only).xlsx")
        self.start_range= datetime(date.today().year, 1, 1).date()
        self.end_range= datetime.now().date()
        self.parsed= parse_excel_to_lists(parse_file, "dailyrate")
        self.plot= generate_plot(self)


def main():
    plot= ClockPlot()
    mission_plot(plot)
    year_plot(plot)

    if input(" - Do you want to open the plot in a web browser? (y/n): ").lower() in ("y"):
        plot.plot.show()

    if input(" - Do you want to create a biannual plot? (y/n): ").lower() in ["y"]:
        biannual_plot(plot)

if __name__ == "__main__":
    main()
