
from PyQt5.QtWidgets import QPushButton, QMessageBox, QFileDialog
from PyQt5.QtCore import Qt
import os.path
from pathlib import Path
# from pathlib import Path
import openpyxl
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from misc import update_svn_button_style


class SVNInstructionBox():
    """
        A custom wrapper for a QMessageBox designed to guide users through 
        the SVN directory selection and file saving process.

        This class provides a modal dialog that informs the user how to 
        correctly map their local SVN checkout for the 'Chandra_Limits' 
        repository before the application attempts to save files.

        Attributes:
            svn_instruction_box (QMessageBox): The underlying message box widget.
            svn_inst_btn (QPushButton): The 'Continue' button assigned the AcceptRole.
    """
    def __init__(self):
        "Defien the SVN Instruction Box"
        self.svn_instruction_box= QMessageBox()
        self.svn_instruction_box.setWindowTitle("SVN Save Steps")
        self.svn_inst_btn= self.svn_instruction_box.addButton(
            "Continue", QMessageBox.ButtonRole.AcceptRole)
        self.svn_instruction_box.addButton(QMessageBox.StandardButton.Cancel)

        instr_text= """
            <div style='margin-left: 10px;'><ol>
                <li><b>Select SVN Directory for "Chandra_Limits":</b><br><br>
                Select where you have checked out the SVN Directory <b>http://svn.occ.harvard.edu/svn/fot/
                Deployment/Tools/Chandra_Limits</b>. If you do not have this checked-out locally
                you must locally check it out.<br><br></li>
                <li><b>"Sheets Master" Saved to Chosen Directory</b>:<br><br>
                The latest version of the "Limits Sheets Masters" file will be saved as an .xlsx file in 
                the chosen directory.<br><br>
                <li>Select <b>Continue</b> once ready to proceed.
                </ol>
            """
        self.svn_instruction_box.setInformativeText(instr_text)
        self.svn_instruction_box.setTextFormat(Qt.RichText)


class SVNCommitBox():
    """
        A confirmation dialog used to authorize an SVN commit action.

        This class serves as a final gateway before the application executes 
        terminal or API commands to commit changes to the SVN repository. 
        It ensures the user is ready to proceed and provides a clear path 
        to abort the operation.

        Attributes:
            svn_commit_box (QMessageBox): The underlying message box widget.
            svn_cont_btn (QPushButton): The 'Continue' button used to trigger 
                the commit logic.
    """
    def __init__(self):
        "Define the SVN Commit Box"
        self.svn_commit_box= QMessageBox()
        self.svn_commit_box.setWindowTitle("SVN Commit")
        self.svn_cont_btn= self.svn_commit_box.addButton("Continue", QMessageBox.ButtonRole.AcceptRole)
        self.svn_commit_box.addButton(QMessageBox.StandardButton.Cancel)

        commit_text= """
            <div style='margin-left: 10px;'>
                Select <b>Continue</b> to proceed with SVN Commit.<br><br>
                Select <b>Cancel</b> to cancel SVN Commit.
            """
        self.svn_commit_box.setInformativeText(commit_text)
        self.svn_commit_box.setTextFormat(Qt.RichText)


def try_numeric(value):
    """
        Attempts to convert a string value into a native Python integer or float.

        This utility is primarily used when preparing data for Excel exports to 
        ensure that numeric strings are stored as numbers rather than text. This 
        enables proper sorting, filtering, and mathematical calculations within 
        the spreadsheet.

        The conversion follows a specific priority:
        1. Integer (if the string represents a whole number).
        2. Float (if the string represents a decimal).
        3. Original value (if conversion fails or the input is not a string).

        Args:
            value (any): The input data to be checked. Usually a string extracted 
                from a UI element or a raw data file.

        Returns:
            int | float | any: The converted numeric value if successful; 
                otherwise, returns the original value unchanged.

        Example:
            >>> try_numeric(" 123 ")
            123
            >>> try_numeric("45.67")
            45.67
            >>> try_numeric("Chandra")
            'Chandra'
    """
    if isinstance(value, str):
        # Remove whitespace to prevent conversion errors
        clean_val = value.strip()
        
        # Try Integer first
        try:
            return int(clean_val)
        except ValueError:
            pass
        
        # Try Float second
        try:
            return float(clean_val)
        except ValueError:
            pass
            
    return value


def sheet_data_to_excel(self):
    """
        Converts retrieved Google Sheets data into a formatted Excel (.xlsx) file.

        This method initializes a new OpenPyXL workbook, applies custom styling 
        (borders, fonts, and fills), and saves the final product to the local 
        SVN directory.

        Process Flow:
        1.  **Initialization**: Creates a workbook and sets the active sheet title.
        2.  **Data Processing**: Iterates through `self.sheets_data`, converting 
            numeric strings to actual integers/floats using `try_numeric`.
        3.  **Styling**: 
            - Applies a gray 'PatternFill' and bold font to the header row.
            - Applies thin borders and center alignment to all cells.
        4.  **Layout Optimization**: Dynamically calculates and adjusts column 
            widths based on the longest string in each column.
        5.  **I/O**: Saves the file as 'chandra_limits.xlsx' in the path defined 
            by `self.svn_path`.

        Raises:
            OSError: If the directory in `self.svn_path` is invalid or unwritable.
            AttributeError: If `self.sheets_data` or `self.svn_path` are not defined.
    """
    wb= openpyxl.Workbook()
    ws= wb.active
    ws.title= "Chandra Limits"

    header_fill= PatternFill(start_color= "D9D9D9", end_color= "D9D9D9", fill_type= "solid")
    header_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Append data to excel sheet
    for element in self.sheets_data:
        converted_row= [try_numeric(item) for item in element]
        ws.append(converted_row)

    # Apply formatting to Header (Row 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Apply formatting to Data (All other rows)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    wb.save(os.path.join(self.svn_path,"chandra_limits.xlsx"))


def open_svn_save_diaglog(self):
    "Open the SVN Directory Select Dialog Box"

    instr_box= SVNInstructionBox()
    commit_box= SVNCommitBox()
    instr_box.svn_instruction_box.exec_()

    if instr_box.svn_instruction_box.clickedButton() == instr_box.svn_inst_btn:
        self.svn_path= QFileDialog.getExistingDirectory(self, "Select SVN Directory")
        sheet_data_to_excel(self)
        QMessageBox.information(self, "File Successfully Saved",
                                f"Successfully saved chandra_limits.xlsx to {self.svn_path}")
        commit_box.svn_commit_box.exec_()

        if commit_box.svn_commit_box.clickedButton() == commit_box.svn_cont_btn:
            print("Call SVN Commit Action here....")


def add_save_to_svn_btn(self):
    "Add the 'Save to SVN' button to the GUI"
    self.btn_svn = QPushButton("Save to SVN")
    self.btn_svn.setEnabled(False)
    self.btn_svn.clicked.connect(lambda: open_svn_save_diaglog(self))
    self.layout.addWidget(self.btn_svn)
    update_svn_button_style(self, enabled=False) # Set initial grey style
