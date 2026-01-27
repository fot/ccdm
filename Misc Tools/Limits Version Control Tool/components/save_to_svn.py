
from PyQt5.QtWidgets import QPushButton, QMessageBox, QFileDialog
from PyQt5.QtCore import Qt
import os.path
import openpyxl
import os
import subprocess
import shutil
import tempfile
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from components.misc import update_svn_button_style
from components.google_auth import get_sheets_api_service


class SVNInstructionBox():
    """
        A custom wrapper for a QMessageBox designed to guide users through 
        the SVN directory selection and file saving process.

        This class provides a modal dialog that informs the user how to 
        correctly map their local SVN checkout for the 'Chandra_Limits' 
        repository before the application attempts to save files.

        Attributes:
            svn_instruction_box (QMessageBox): The underlying message box widget.
            svn_inst_btn (QPushButton): The 'Continue' button assigned the AcceptRole.
    """
    def __init__(self):
        "Defien the SVN Instruction Box"
        self.svn_instruction_box= QMessageBox()
        self.svn_instruction_box.setWindowTitle("SVN Save Steps")
        self.svn_inst_btn= self.svn_instruction_box.addButton(
            "Continue", QMessageBox.ButtonRole.AcceptRole)
        self.svn_instruction_box.addButton(QMessageBox.StandardButton.Cancel)

        instr_text= """
            <div style='margin-left: 10px;'><ol>
                <li><b>Select SVN Directory for "Chandra_Limits":</b><br><br>
                Select where you have checked out the SVN Directory <b>http://svn.occ.harvard.edu/svn/fot/
                Deployment/Tools/Chandra_Limits</b>. If you do not have this checked-out locally
                you must locally check it out.<br><br></li>
                <li><b>"Sheets Master" Saved to Chosen Directory</b>:<br><br>
                The latest version of the "Limits Sheets Masters" file will be saved as an .xlsx file in 
                the chosen directory.<br><br>
                <li>Select <b>Continue</b> once ready to proceed.
                </ol>
            """
        self.svn_instruction_box.setInformativeText(instr_text)
        self.svn_instruction_box.setTextFormat(Qt.RichText)


class SVNCommitBox():
    """
        A confirmation dialog used to authorize an SVN commit action.

        This class serves as a final gateway before the application executes 
        terminal or API commands to commit changes to the SVN repository. 
        It ensures the user is ready to proceed and provides a clear path 
        to abort the operation.

        Attributes:
            svn_commit_box (QMessageBox): The underlying message box widget.
            svn_cont_btn (QPushButton): The 'Continue' button used to trigger 
                the commit logic.
    """
    def __init__(self):
        "Define the SVN Commit Box"
        self.svn_commit_box= QMessageBox()
        self.svn_commit_box.setWindowTitle("SVN Commit")
        self.svn_cont_btn= self.svn_commit_box.addButton("Continue", QMessageBox.ButtonRole.AcceptRole)
        self.svn_commit_box.addButton(QMessageBox.StandardButton.Cancel)

        commit_text= """
            <div style='margin-left: 10px;'>
                Select <b>Continue</b> to proceed with SVN Commit.<br><br>
                Select <b>Cancel</b> to cancel SVN Commit.
            """
        self.svn_commit_box.setInformativeText(commit_text)
        self.svn_commit_box.setTextFormat(Qt.RichText)


def try_numeric(value):
    """
        Attempts to convert a string value into a native Python integer or float.

        This utility is primarily used when preparing data for Excel exports to 
        ensure that numeric strings are stored as numbers rather than text. This 
        enables proper sorting, filtering, and mathematical calculations within 
        the spreadsheet.

        The conversion follows a specific priority:
        1. Integer (if the string represents a whole number).
        2. Float (if the string represents a decimal).
        3. Original value (if conversion fails or the input is not a string).

        Args:
            value (any): The input data to be checked. Usually a string extracted 
                from a UI element or a raw data file.

        Returns:
            int | float | any: The converted numeric value if successful; 
                otherwise, returns the original value unchanged.

        Example:
            >>> try_numeric(" 123 ")
            123
            >>> try_numeric("45.67")
            45.67
            >>> try_numeric("Chandra")
            'Chandra'
    """
    if isinstance(value, str):
        # Remove whitespace to prevent conversion errors
        clean_val = value.strip()
        
        # Try Integer first
        try:
            return int(clean_val)
        except ValueError:
            pass
        
        # Try Float second
        try:
            return float(clean_val)
        except ValueError:
            pass
            
    return value


def sheet_data_to_excel(self):
    """
        Converts retrieved Google Sheets data into a formatted Excel (.xlsx) file.

        This method initializes a new OpenPyXL workbook, applies custom styling 
        (borders, fonts, and fills), and saves the final product to the local 
        SVN directory.

        Process Flow:
        1.  **Initialization**: Creates a workbook and sets the active sheet title.
        2.  **Data Processing**: Iterates through `self.sheets_data`, converting 
            numeric strings to actual integers/floats using `try_numeric`.
        3.  **Styling**: 
            - Applies a gray 'PatternFill' and bold font to the header row.
            - Applies thin borders and center alignment to all cells.
        4.  **Layout Optimization**: Dynamically calculates and adjusts column 
            widths based on the longest string in each column.
        5.  **I/O**: Saves the file as 'chandra_limits.xlsx' in the path defined 
            by `self.svn_path`.

        Raises:
            OSError: If the directory in `self.svn_path` is invalid or unwritable.
            AttributeError: If `self.sheets_data` or `self.svn_path` are not defined.
    """

    # Refresh the sheets data in the event it was edited externally
    _, self.sheets_data= get_sheets_api_service(self)

    wb= openpyxl.Workbook()
    ws= wb.active
    ws.title= "Chandra Limits"

    header_fill= PatternFill(start_color= "D9D9D9", end_color= "D9D9D9", fill_type= "solid")
    header_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Append data to excel sheet
    for element in self.sheets_data:
        converted_row= [try_numeric(item) for item in element]
        ws.append(converted_row)

    # Apply formatting to Header (Row 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Apply formatting to Data (All other rows)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    file_output_path= Path(self.svn_path) / "chandra_limits.xlsx"

    try:
        wb.save(file_output_path)
    except PermissionError:
        # This triggers if the file is open in Excel
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setText("Cannot Save File")
        msg.setInformativeText(
            f"The file '{file_output_path.name}' is currently open in another program (likely Excel).\n\n"
            "Please close the file and click OK to try again."
        )
        msg.setWindowTitle("Permission Error")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        
        if msg.exec_() == QMessageBox.Ok:
            # Recursive call to try saving again
            sheet_data_to_excel(self) 
        else:
            return False # User cancelled


def svn_commit_file(self):
    """
        Commits the 'chandra_limits.xlsx' file to the Subversion repository using 
        bundled portable SVN tools.

        This function handles the specific challenges of running portable binaries 
        from a network share (UNC paths) and managing DLL dependencies in 
        modern Python environments.

        Key Features:
        ----------
        - Dynamic Path Resolution: Locates 'svn-tools' relative to the script 
        location to maintain portability across different user environments.
        - DLL Management: Uses 'os.add_dll_directory' (Python 3.8+) to explicitly 
        trust the bundled bin folder, resolving 'libapr-1.dll not found' errors.
        - Path Priority: Prepend the bin directory to the system PATH environment 
        variable for the duration of the subprocess call.
        - UNC Path Workaround: Sets the 'cwd' (Current Working Directory) to a 
        local disk path (the file's parent directory) to prevent CMD errors 
        when executing binaries from a network share.

        Attributes Required:
        -------------------
        - self.svn_path (str/Path): The local directory of the SVN checkout.
        - self.ticket_obj (str, optional): JIRA ticket number for the commit message.

        Returns:
        -------
        - bool: True if 'add' and 'commit' operations were successful, False otherwise.

        Raises:
        -------
        - subprocess.CalledProcessError: Captured internally; prints STDOUT and 
        STDERR to the console for debugging.
    """
    # Locate the svn-tools directory relative to this component file
    component_dir = Path(__file__).resolve().parent

    # Try current folder, then one level up (root)
    svn_bin_dir = component_dir / "svn-tools" / "bin"
    if not svn_bin_dir.exists():
        svn_bin_dir = component_dir.parent / "svn-tools" / "bin"

    svn_exe = svn_bin_dir / "svn.exe"

    # Safety Check
    if not svn_exe.exists():
        print(f"CRITICAL: svn.exe not found at: {svn_exe}")
        return False

    # Pathing for Excel file
    file_path = Path(self.svn_path) / "chandra_limits.xlsx"
    ticket = getattr(self, "ticket_obj", "No JIRA Ticket")
    commit_message = f"Updated chandra_limits.xlsx ({ticket})"

    # Environment and DLL Handling
    # Adding the bin directory to PATH ensures svn.exe can see libapr-1.dll 
    # even when running from a network share.
    env = os.environ.copy()
    env["PATH"] = str(svn_bin_dir) + os.pathsep + env.get("PATH", "")

    # For Python 3.8+, we must explicitly trust the DLL directory
    dll_handler = None
    if hasattr(os, 'add_dll_directory'):
        dll_handler = os.add_dll_directory(str(svn_bin_dir))

    try:
        subprocess.run(
            [str(svn_exe), "add", str(file_path), "--force"],
            check=True,
            capture_output=True,
            text=True,
            env=env,
            cwd=str(file_path.parent) # Change working directory to local disk
        )

        subprocess.run(
            [str(svn_exe), "commit", str(file_path), "-m", commit_message],
            check=True,
            capture_output=True,
            text=True,
            env=env,
            cwd=str(file_path.parent)
        )
        return True

    except subprocess.CalledProcessError as e:
        print(f"SVN FAILED.\nSTDOUT: {e.stdout}\nSTDERR: {e.stderr}")
        return False

    finally:
        if dll_handler:
            dll_handler.close()


def open_svn_save_diaglog(self):
    "Open the SVN Directory Select Dialog Box"
    instr_box= SVNInstructionBox()
    commit_box= SVNCommitBox()
    instr_box.svn_instruction_box.exec_()

    if instr_box.svn_instruction_box.clickedButton() == instr_box.svn_inst_btn:
        self.svn_path= QFileDialog.getExistingDirectory(self, "Select SVN Directory")
        sheet_data_to_excel(self)
        QMessageBox.information(self, "File Successfully Saved",
                                f"Successfully saved chandra_limits.xlsx to {self.svn_path}")
        commit_box.svn_commit_box.exec_()

        if commit_box.svn_commit_box.clickedButton() == commit_box.svn_cont_btn:
            commit_success= svn_commit_file(self)

            if commit_success:
                QMessageBox.information(self, "File Successfully Commited",
                                        f"Successfully SVN commited chandra_limits.xlsx to repo")
            else:
                QMessageBox.information(self, "File Commit Failed",
                                        f"Diddy didn't do it.")


def add_save_to_svn_btn(self):
    "Add the 'Save to SVN' button to the GUI"
    self.btn_svn = QPushButton("Save to SVN")
    self.btn_svn.setEnabled(False)
    self.btn_svn.clicked.connect(lambda: open_svn_save_diaglog(self))
    self.layout.addWidget(self.btn_svn)
    update_svn_button_style(self, enabled=False) # Set initial grey style
