"Functions to merge input excel file data to master"

from pathlib import Path
from PyQt5.QtWidgets import QMessageBox, QPushButton, QHBoxLayout
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import QUrl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from components.google_auth import get_sheets_api_service
from components.load_file import clear_loaded_file
from components.misc import create_separator


SPREADSHEET_ID= "15rRk5JAMWXBGiKTly4aP0cUuFE1qECZe01tNESSKXBo"


def update_google_sheet(self):
    "update the google sheets file via API"

    body= {"values": self.excel_file_data[1:]}

    self.sheets_service.spreadsheets().values().append(
        spreadsheetId= SPREADSHEET_ID,
        range= "Sheet1!A1:P",
        valueInputOption= "USER_ENTERED",
        body= body
        ).execute()


def update_sheet_format(self):
    "Update the sheet's format after updating."
    spreadsheet = self.sheets_service.spreadsheets().get(spreadsheetId= SPREADSHEET_ID).execute()
    sheet_id = next(s['properties']['sheetId'] for s in spreadsheet['sheets'] 
                        if s['properties']['title'] == "Sheet1")
    num_rows= len(self.sheets_data) + len(self.excel_file_data) - 1
    num_cols= 17  # Columns A through Q

    border_style = {
        "style": "SOLID",
        "width": 1,
        "color": {"red": 0, "green": 0, "blue": 0}
    }

    requests = [
        # 1. Header Row (Exactly Row 0, Columns 0-15)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": num_cols
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.85, "green": 0.85, "blue": 0.85},
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "CENTER",
                        "borders": {
                            "top": border_style, "bottom": border_style, 
                            "left": border_style, "right": border_style
                        }
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,borders)"
            }
        },
        # 2. Data Rows (From Row 1 to num_rows, Columns 0-15)
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": num_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": num_cols
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                        "borders": {
                            "top": border_style, "bottom": border_style, 
                            "left": border_style, "right": border_style
                        }
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment,borders)"
            }
        }
    ]

    self.sheets_service.spreadsheets().batchUpdate(spreadsheetId= SPREADSHEET_ID, body={"requests": requests}).execute()


def append_jira_ticket_num(self):
    "Append the jira ticket number to the end of the excel sheet data before push to google sheets"
    for row in self.excel_file_data:
        try:
            row[16] = str(self.ticket_obj)
        except IndexError:
            row.append(str(self.ticket_obj))


def try_numeric(value):
    """
        Attempts to convert a string value into a native Python integer or float.

        This utility is primarily used when preparing data for Excel exports to 
        ensure that numeric strings are stored as numbers rather than text. This 
        enables proper sorting, filtering, and mathematical calculations within 
        the spreadsheet.

        The conversion follows a specific priority:
        1. Integer (if the string represents a whole number).
        2. Float (if the string represents a decimal).
        3. Original value (if conversion fails or the input is not a string).

        Args:
            value (any): The input data to be checked. Usually a string extracted 
                from a UI element or a raw data file.

        Returns:
            int | float | any: The converted numeric value if successful; 
                otherwise, returns the original value unchanged.

        Example:
            >>> try_numeric(" 123 ")
            123
            >>> try_numeric("45.67")
            45.67
            >>> try_numeric("Chandra")
            'Chandra'
    """
    if isinstance(value, str):
        # Remove whitespace to prevent conversion errors
        clean_val = value.strip()
        
        # Try Integer first
        try:
            return int(clean_val)
        except ValueError:
            pass
        
        # Try Float second
        try:
            return float(clean_val)
        except ValueError:
            pass
            
    return value


def sheet_data_to_excel(self, to_svn= False):
    """
        Converts retrieved Google Sheets data into a formatted Excel (.xlsx) file.

        This method initializes a new OpenPyXL workbook, applies custom styling 
        (borders, fonts, and fills), and saves the final product to the local 
        SVN directory.

        Process Flow:
        1.  **Initialization**: Creates a workbook and sets the active sheet title.
        2.  **Data Processing**: Iterates through `self.sheets_data`, converting 
            numeric strings to actual integers/floats using `try_numeric`.
        3.  **Styling**: 
            - Applies a gray 'PatternFill' and bold font to the header row.
            - Applies thin borders and center alignment to all cells.
        4.  **Layout Optimization**: Dynamically calculates and adjusts column 
            widths based on the longest string in each column.
        5.  **I/O**: Saves the file as 'chandra_limits.xlsx' in the path defined 
            by `self.svn_path`.

        Raises:
            OSError: If the directory in `self.svn_path` is invalid or unwritable.
            AttributeError: If `self.sheets_data` or `self.svn_path` are not defined.
    """

    # Refresh the sheets data in the event it was edited externally
    _, self.sheets_data= get_sheets_api_service(self)

    wb= openpyxl.Workbook()
    ws= wb.active
    ws.title= "Chandra Limits"

    if to_svn:
        file_output_path= Path(self.svn_path) / "chandra_limits.xlsx"
    else:
        file_output_path= Path.home() / "Desktop" / "chandra_limits.xlsx"

    header_fill= PatternFill(start_color= "D9D9D9", end_color= "D9D9D9", fill_type= "solid")
    header_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center", vertical="center")

    # Append data to excel sheet
    for element in self.sheets_data:
        converted_row= [try_numeric(item) for item in element]
        ws.append(converted_row)

    # Apply formatting to Header (Row 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Apply formatting to Data (All other rows)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    try:
        wb.save(file_output_path)
        save_msg= QMessageBox()
        save_msg.setWindowTitle("File Saved...")
        save_msg.setText("Successfully saved Google Sheets Master data "
                         f"to chandra_limits.xlsx! on {file_output_path.parent}.")
        save_msg.setIcon(QMessageBox.Icon.Information)
        save_msg.exec()
    except PermissionError:
        # This triggers if the file is open in Excel
        err_msg = QMessageBox()
        err_msg.setIcon(QMessageBox.Critical)
        err_msg.setText("Cannot Save File")
        err_msg.setInformativeText(
            f"The file '{file_output_path.name}' is currently open in another program (likely Excel).\n\n"
            "Please close the file and click OK to try again."
        )
        err_msg.setWindowTitle("Permission Error")
        err_msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        
        if err_msg.exec_() == QMessageBox.Ok:
            # Recursive call to try saving again
            sheet_data_to_excel(self)


def merge_to_master(self):
    "Actions when the 'Merge to Master' button is clicked"
    
    mrg_msg= QMessageBox(self)
    mrg_msg.setWindowTitle("Merge to Google Sheet Master")
    mrg_msg.setText(f"Proceed with merging data from ({self.fileName}) to Master Google Sheets File?")
    mrg_msg.setIcon(QMessageBox.Icon.Information)

    # Add buttons and change the text of the "Ok" button
    continue_button = mrg_msg.addButton("Continue", QMessageBox.ButtonRole.AcceptRole)
    mrg_msg.addButton(QMessageBox.StandardButton.Cancel)
    mrg_msg.exec()

    if mrg_msg.clickedButton() == continue_button:

        # Refresh the sheets data in the event it was edited externally
        _, self.sheets_data= get_sheets_api_service(self)

        # Update Google Sheets file with excel data
        append_jira_ticket_num(self)
        update_google_sheet(self)
        update_sheet_format(self)
        QMessageBox.information(self, "Success",
                        f"Successfully merged spreadsheet ({self.fileName}) to Google Sheets Master!")
        clear_loaded_file(self)


def open_sheets_master():
    url = QUrl("https://docs.google.com/spreadsheets/d/15rRk5JAMWXBGiKTly4aP0cUuFE1qECZe01tNESSKXBo/edit?usp=sharing")
    QDesktopServices.openUrl(url)


def add_sheet_master_btns(self):
    "Add the 'Merge to Master' button to the GUI"
    self.merge_btn= QPushButton("Merge to Master")
    self.merge_btn.clicked.connect(lambda: merge_to_master(self))
    self.save_btn= QPushButton("Save Master Locally")
    self.save_btn.clicked.connect(lambda: sheet_data_to_excel(self))
    self.open_master_btn= QPushButton("Open Sheets Master")
    self.open_master_btn.clicked.connect(lambda: open_sheets_master())
    self.open_master_btn.setStyleSheet("background-color: #28a745; color: white; font-size: 12px;"
                                       "font-weight: bold; padding: 12px; border-radius: 5px;")
    self.master_btn_layout= QHBoxLayout()
    self.master_btn_layout.addWidget(self.merge_btn)
    self.master_btn_layout.addWidget(self.save_btn)
    self.layout.addLayout(self.master_btn_layout)
    self.layout.addWidget(self.open_master_btn)
    self.layout.addWidget(create_separator(self))
